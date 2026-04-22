from __future__ import annotations

import argparse
import importlib
import math
import re
import time
import unicodedata
import json
import os
import random
import requests
from collections import Counter
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

DATA_PATH = Path(__file__).with_name("tmdb_top1000_movies.xlsx")
DESCRIPTION_LIMIT = 500
MODEL = "gemma4:31b-cloud"
OLLAMA_HOST = "https://ollama.com"
LLM_CHAR_BUDGET = 420
REQUEST_TIMEOUT_SECONDS = 8
CRITIC_REVIEW_MIN_BUDGET_SECONDS = 1.6
CRITIC_REWRITE_MIN_BUDGET_SECONDS = 1.0
CRITIC_REVIEW_TIMEOUT_SECONDS = 1.2
CRITIC_REWRITE_TIMEOUT_SECONDS = 1.2
SEMANTIC_RERANK_POOL_SIZE = 30
SEMANTIC_SCORE_WEIGHT = 8.0

STOPWORDS = {
    "a", "about", "after", "all", "an", "and", "any", "are", "as", "at", "be",
    "because", "but", "by", "for", "from", "have", "i", "if", "im", "in", "into",
    "is", "it", "its", "just", "kind", "like", "me", "movie", "movies", "my",
    "of", "on", "or", "something", "that", "the", "them", "these", "those", "to",
    "want", "with",
}

EXCLUSION_STOPWORDS = STOPWORDS | {
    "not", "ones", "one", "movies", "movie", "films", "film", "stuff", "anything"
}

GENRE_ALIASES = {
    "action": {"action", "adrenaline", "explosive", "fight", "fight scenes"},
    "adventure": {"adventure", "quest", "journey", "epic"},
    "animation": {"animated", "animation", "cartoon", "pixar", "disney"},
    "comedy": {"comedy", "comedic", "funny", "humor", "laughs", "witty"},
    "crime": {"crime", "gangster", "heist", "mafia", "mob"},
    "drama": {"drama", "dramatic", "character study", "emotional"},
    "family": {"family", "kids", "kid", "all ages", "family night"},
    "fantasy": {"fantasy", "magic", "magical", "mythic"},
    "history": {"historical", "history", "period piece"},
    "horror": {"horror", "scary", "creepy", "terrifying", "haunted"},
    "music": {"music", "musical", "songs", "singing"},
    "mystery": {"mystery", "whodunit", "investigation", "twist"},
    "romance": {"romance", "romantic", "rom-com", "romcom", "love story"},
    "science fiction": {"sci-fi", "science fiction", "space", "futuristic"},
    "superhero": {"superhero", "superheroes", "marvel", "mcu", "dc comics", "comic book"},
    "thriller": {"thriller", "tense", "suspense", "suspenseful"},
    "war": {"war", "military", "battlefront"},
}

LANGUAGE_ALIASES = {
    "english": {"en"},
    "korean": {"ko"},
    "japanese": {"ja"},
    "french": {"fr"},
    "spanish": {"es"},
    "italian": {"it"},
    "german": {"de"},
    "hindi": {"hi"},
    "portuguese": {"pt"},
    "chinese": {"zh"},
    "mandarin": {"zh"},
}

COUNTRY_ALIASES = {
    "brazil": {"country_tokens": {"brazil"}, "language_codes": {"pt"}},
    "brazilian": {"country_tokens": {"brazil"}, "language_codes": {"pt"}},
    "portugal": {"country_tokens": {"portugal"}, "language_codes": {"pt"}},
    "portuguese": {"country_tokens": {"portugal", "brazil"}, "language_codes": {"pt"}},
    "china": {"country_tokens": {"china"}, "language_codes": {"zh"}},
    "chinese": {"country_tokens": {"china"}, "language_codes": {"zh"}},
    "japan": {"country_tokens": {"japan"}, "language_codes": {"ja"}},
    "japanese": {"country_tokens": {"japan"}, "language_codes": {"ja"}},
    "korea": {"country_tokens": {"korea"}, "language_codes": {"ko"}},
    "korean": {"country_tokens": {"korea"}, "language_codes": {"ko"}},
    "france": {"country_tokens": {"france"}, "language_codes": {"fr"}},
    "french": {"country_tokens": {"france"}, "language_codes": {"fr"}},
    "italy": {"country_tokens": {"italy"}, "language_codes": {"it"}},
    "italian": {"country_tokens": {"italy"}, "language_codes": {"it"}},
    "germany": {"country_tokens": {"germany"}, "language_codes": {"de"}},
    "german": {"country_tokens": {"germany"}, "language_codes": {"de"}},
    "spain": {"country_tokens": {"spain"}, "language_codes": {"es"}},
    "spanish": {"country_tokens": {"spain"}, "language_codes": {"es"}},
    "mexico": {"country_tokens": {"mexico"}, "language_codes": {"es"}},
    "mexican": {"country_tokens": {"mexico"}, "language_codes": {"es"}},
    "argentina": {"country_tokens": {"argentina"}, "language_codes": {"es"}},
    "argentinian": {"country_tokens": {"argentina"}, "language_codes": {"es"}},
    "argentine": {"country_tokens": {"argentina"}, "language_codes": {"es"}},
    "india": {"country_tokens": {"india"}, "language_codes": {"hi"}},
    "indian": {"country_tokens": {"india"}, "language_codes": {"hi"}},
}

STUDIO_ALIASES = {
    "pixar": {"company_tokens": {"pixar"}, "preferred_genres": {"animation", "family"}, "language_codes": {"en"}},
    "disney": {"company_tokens": {"disney"}, "preferred_genres": {"animation", "family"}, "language_codes": {"en"}},
    "studio ghibli": {"company_tokens": {"ghibli"}, "preferred_genres": {"animation", "family", "fantasy"}, "language_codes": {"ja"}},
    "ghibli": {"company_tokens": {"ghibli"}, "preferred_genres": {"animation", "family", "fantasy"}, "language_codes": {"ja"}},
    "warner bros": {"company_tokens": {"warner", "bros"}, "preferred_genres": set(), "language_codes": set()},
    "warner brothers": {"company_tokens": {"warner", "bros"}, "preferred_genres": set(), "language_codes": set()},
    "a24": {"company_tokens": {"a24"}, "preferred_genres": set(), "language_codes": set()},
    "marvel studios": {"company_tokens": {"marvel"}, "preferred_genres": {"superhero", "action"}, "language_codes": set()},
    "20th century fox": {"company_tokens": {"20th", "century", "fox"}, "preferred_genres": set(), "language_codes": set()},
    "fox searchlight": {"company_tokens": {"fox", "searchlight"}, "preferred_genres": set(), "language_codes": set()},
    "searchlight": {"company_tokens": {"searchlight"}, "preferred_genres": set(), "language_codes": set()},
    "universal": {"company_tokens": {"universal"}, "preferred_genres": set(), "language_codes": set()},
    "universal pictures": {"company_tokens": {"universal"}, "preferred_genres": set(), "language_codes": set()},
    "paramount": {"company_tokens": {"paramount"}, "preferred_genres": set(), "language_codes": set()},
    "paramount pictures": {"company_tokens": {"paramount"}, "preferred_genres": set(), "language_codes": set()},
}

EXCLUSION_HINTS = {
    "marvel": {"company_tokens": {"marvel", "mcu"}, "genre_tokens": {"superhero"}},
    "mcu": {"company_tokens": {"marvel", "mcu"}, "genre_tokens": {"superhero"}},
    "dc": {"company_tokens": {"dc"}, "genre_tokens": {"superhero"}},
    "dc comics": {"company_tokens": {"dc"}, "genre_tokens": {"superhero"}},
    "animated": {"company_tokens": set(), "genre_tokens": {"animation"}},
    "animation": {"company_tokens": set(), "genre_tokens": {"animation"}},
}

TV_SERIES_REQUEST_PHRASES = {
    "tv series",
    "television series",
    "tv show",
    "television show",
}

TV_TIE_IN_KEYWORD_PHRASES = {
    "based on tv series",
    "edited from tv series",
    "based on cartoon",
}

PHRASE_HINTS = {
    "superhero": ["superhero", "comic", "hero", "villain", "marvel", "dc"],
    "buddy cop": ["buddy", "banter", "team-up", "partners", "crime", "cop"],
    "feel good": ["uplifting", "warm", "fun", "heart", "hopeful"],
    "emotional": ["moving", "heartwarming", "character", "drama"],
    "strong story": ["drama", "character", "relationship", "psychological", "mystery"],
    "interesting characters": ["character", "relationship", "ensemble", "psychological", "drama"],
    "character driven": ["character", "drama", "relationship", "psychological"],
    "heart": ["heartwarming", "warm", "hopeful"],
    "twist ending": ["twist", "reveal", "mystery", "psychological", "mind-bending"],
    "slow burn": ["atmospheric", "patient", "moody", "character", "tension"],
    "exciting": ["action", "thriller", "intense", "momentum"],
    "not mindless": ["smart", "clever", "character", "driven"],
    "dread": ["psychological", "atmospheric", "tension"],
    "chemistry": ["romance", "spark", "relationship"],
    "unreliable": ["psychological", "paranoia", "mystery"],
    "family movie night": ["family", "heartwarming", "adventure", "fun", "crowd-pleasing"],
    "rom com": ["romance", "chemistry", "charming", "comedy"],
    "mind bending": ["mind-bending", "dream", "reality", "psychological", "mystery"],
}

PIXAR_HINT_TOKENS = ("family", "heartwarming", "friendship", "adventure", "comedy", "hopeful")

HAPPY_ENDING_AVOIDANCE_PHRASES = {
    "hate sad ending",
    "hate sad endings",
    "hates sad ending",
    "hates sad endings",
    "not sad",
    "not too sad",
    "isnt sad",
    "nothing sad",
    "no sad ending",
    "no sad endings",
    "no sad movie",
    "no sad movies",
    "without sad ending",
    "without sad endings",
    "happy ending",
    "happy endings",
    "uplifting ending",
    "uplifting endings",
}

SAD_THEME_TOKENS = {
    "grief",
    "tragic",
    "tragedy",
    "mourning",
    "melancholy",
    "loss",
    "bleak",
    "heartbreak",
    "devastating",
    "sacrifice",
}

HAPPY_THEME_TOKENS = {
    "friendship",
    "family",
    "hope",
    "hopeful",
    "heartwarming",
    "uplifting",
    "joy",
    "adventure",
}

LOW_RATED_PHRASES = {
    "rated poorly",
    "poorly rated",
    "low rated",
    "low-rated",
    "bad movie",
    "bad movies",
    "bad film",
    "bad films",
    "terrible movie",
    "terrible movies",
    "awful movie",
    "awful movies",
    "bad on purpose",
    "so bad it is good",
    "so bad it s good",
    "so bad its good",
    "so bad it's good",
    "disasterpiece",
    "campy",
    "trashy",
    "cheesy",
    "hate-watch",
    "hate watch",
    "trainwreck",
    "worst movie",
    "worst movies",
}

ABSOLUTE_WORST_PHRASES = {
    "worst movie ever",
    "worst movie ever made",
    "the worst movie ever",
    "the worst movie ever made",
    "absolute worst movie",
    "single worst movie",
    "lowest rated movie",
    "lowest-rated movie",
}

STYLE_PROFILES = [
    {
        "instruction": "Write in a cinematic, vivid, high-momentum voice.",
        "hook_prefix": "If you want something that feels big right away,",
        "fit_verb": "unfolds through",
        "payoff_prefix": "What really lands is that",
    },
    {
        "instruction": "Write like a witty friend making a sharp, playful recommendation.",
        "hook_prefix": "If you want a pick with some personality,",
        "fit_verb": "clicks through",
        "payoff_prefix": "What makes it fun is that",
    },
    {
        "instruction": "Write in a crisp, direct, no-fluff voice.",
        "hook_prefix": "If you want a clear, strong pick,",
        "fit_verb": "works through",
        "payoff_prefix": "The payoff is that",
    },
    {
        "instruction": "Write in a warm, conversational, natural voice.",
        "hook_prefix": "If you want something easy to get into,",
        "fit_verb": "draws you in through",
        "payoff_prefix": "What lands nicely is that",
    },
    {
        "instruction": "Write like an accessible but sharp film critic.",
        "hook_prefix": "If you want something with a real point of view,",
        "fit_verb": "stands out through",
        "payoff_prefix": "What sells it is that",
    },
]

LOW_RATED_STYLE_PROFILES = [
    {
        "instruction": "Write like you're pitching a gloriously messy, campy bad-movie night.",
        "hook_prefix": "If you want something gloriously messy,",
        "fit_verb": "leans into",
        "payoff_prefix": "The fun of it is that",
    },
    {
        "instruction": "Write like you're recommending a chaotic hate-watch with real enthusiasm.",
        "hook_prefix": "If you want a full-on chaotic pick,",
        "fit_verb": "runs on",
        "payoff_prefix": "The appeal here is that",
    },
    {
        "instruction": "Write like a friend recommending an intentionally rough, weirdly watchable trainwreck.",
        "hook_prefix": "If you want a bad-movie night on purpose,",
        "fit_verb": "coasts on",
        "payoff_prefix": "What makes it work as a bad-movie pick is that",
    },
]

NEGATION_PHRASES = [
    "no ", "not ", "avoid ", "hate ", "hates ", "dislike ", "dislikes ", "dont want ", "don't want ", "don t want ", "without ",
    "dont like ", "don't like ", "don t like ", "do not like ",
    "except ", "anything except ", "other than ",
]


@dataclass(frozen=True)
class Movie:
    tmdb_id: int
    title: str
    original_language: str
    year: int | None
    runtime_min: int | None
    genres: tuple[str, ...]
    production_companies: tuple[str, ...]
    production_countries: tuple[str, ...]
    overview: str
    tagline: str
    director: str
    cast: tuple[str, ...]
    keywords: tuple[str, ...]
    popularity: float
    vote_average: float
    vote_count: int
    us_rating: str
    normalized_title: str
    title_key: str
    token_set: frozenset[str]
    production_company_tokens: frozenset[str]
    production_country_tokens: frozenset[str]
    searchable_blob: str
    quality_score: float


def _ascii_text(text: str) -> str:
    return unicodedata.normalize("NFKD", text or "").encode("ascii", "ignore").decode("ascii")


def normalize_text(text: str) -> str:
    lowered = _ascii_text(text).lower()
    lowered = re.sub(r"&", " and ", lowered)
    lowered = re.sub(r"[^a-z0-9]+", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def title_key(text: str) -> str:
    cleaned = normalize_text(text)
    cleaned = re.sub(r"\b(the|a|an)\b\s*", "", cleaned)
    cleaned = re.sub(r"\b\d{4}\b", "", cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def tokenize(text: str) -> list[str]:
    return [token for token in normalize_text(text).split() if len(token) > 2 and token not in STOPWORDS]


def contains_normalized_phrase(normalized_text: str, phrase: str) -> bool:
    phrase_norm = normalize_text(phrase)
    if not phrase_norm:
        return False
    return f" {phrase_norm} " in f" {normalized_text} "


def extract_runtime_limit_minutes(text: str, normalized_text: str) -> int | None:
    raw_text = re.sub(r"\s+", " ", _ascii_text(text).lower()).strip()
    if not any(unit in raw_text for unit in ["hour", "hours", "hr", "hrs", "minute", "minutes", "min"]):
        return None

    budget_cues = [
        "only have",
        "just have",
        "have only",
        "only got",
        "got only",
        "time for",
        "fit in",
        "within",
        "under",
        "less than",
        "at most",
        "no more than",
    ]
    if not any(cue in normalized_text for cue in budget_cues):
        return None

    half_hour_match = re.search(r"\b(\d+)\s+and\s+a\s+half\s+hours?\b", raw_text)
    if half_hour_match:
        return (int(half_hour_match.group(1)) * 60) + 30

    hour_and_half_match = re.search(r"\b(?:an?|one)\s+hour\s+and\s+a\s+half\b", raw_text)
    if hour_and_half_match:
        return 90

    minutes_match = re.search(r"\b(\d{2,3})\s*(?:minutes|minute|min)\b", raw_text)
    if minutes_match:
        return int(minutes_match.group(1))

    hours_match = re.search(r"\b(\d+(?:\.\d+)?)\s*(?:hours|hour|hrs|hr)\b", raw_text)
    if hours_match:
        return int(math.floor(float(hours_match.group(1)) * 60))

    return None


def split_csvish(text: str | None) -> tuple[str, ...]:
    if not text:
        return ()
    return tuple(part.strip() for part in str(text).split(",") if part and part.strip())


def _safe_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _safe_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


@lru_cache(maxsize=1)
def load_movies() -> tuple[Movie, ...]:
    workbook = load_workbook(DATA_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    index = {name: idx for idx, name in enumerate(headers)}

    raw_movies: list[dict[str, object]] = []
    max_popularity = 1.0
    max_vote = 1.0
    max_vote_count = 1.0

    for row in rows:
        tmdb_id = _safe_int(row[index["tmdb_id"]])
        title = str(row[index["title"]] or "").strip()
        if not tmdb_id or not title:
            continue

        popularity = _safe_float(row[index["popularity"]])
        vote_average = _safe_float(row[index["vote_average"]])
        vote_count = _safe_int(row[index["vote_count"]]) or 0

        raw_movie = {
            "tmdb_id": tmdb_id,
            "title": title,
            "original_language": str(row[index["original_language"]] or "").strip().lower(),
            "year": _safe_int(row[index["year"]]),
            "runtime_min": _safe_int(row[index["runtime_min"]]),
            "genres": split_csvish(row[index["genres"]]),
            "production_companies": split_csvish(row[index["production_companies"]]),
            "production_countries": split_csvish(row[index["production_countries"]]),
            "overview": str(row[index["overview"]] or "").strip(),
            "tagline": str(row[index["tagline"]] or "").strip(),
            "director": str(row[index["director"]] or "").strip(),
            "cast": split_csvish(row[index["top_cast"]]),
            "keywords": split_csvish(row[index["keywords"]]),
            "popularity": popularity,
            "vote_average": vote_average,
            "vote_count": vote_count,
            "us_rating": str(row[index["us_rating"]] or "").strip(),
        }

        raw_movies.append(raw_movie)
        max_popularity = max(max_popularity, popularity)
        max_vote = max(max_vote, vote_average)
        max_vote_count = max(max_vote_count, float(vote_count))

    movies: list[Movie] = []
    for raw_movie in raw_movies:
        joined = " ".join(
            [
                str(raw_movie["title"]),
                " ".join(raw_movie["genres"]),
                " ".join(raw_movie["production_companies"]),
                " ".join(raw_movie["production_countries"]),
                str(raw_movie["overview"]),
                str(raw_movie["tagline"]),
                str(raw_movie["director"]),
                " ".join(raw_movie["cast"]),
                " ".join(raw_movie["keywords"]),
            ]
        )
        tokens = frozenset(tokenize(joined))
        production_company_tokens = frozenset(
            token
            for company in raw_movie["production_companies"]
            for token in tokenize(company)
        )
        production_country_tokens = frozenset(
            token
            for country in raw_movie["production_countries"]
            for token in tokenize(country)
        )
        popularity = float(raw_movie["popularity"])
        vote_average = float(raw_movie["vote_average"])
        vote_count = int(raw_movie["vote_count"])

        quality_score = (
            0.5 * (vote_average / max_vote)
            + 0.2 * math.log1p(popularity) / math.log1p(max_popularity)
            + 0.3 * math.log1p(vote_count) / math.log1p(max_vote_count)
        )

        movies.append(
            Movie(
                tmdb_id=int(raw_movie["tmdb_id"]),
                title=str(raw_movie["title"]),
                original_language=str(raw_movie["original_language"]),
                year=raw_movie["year"],
                runtime_min=raw_movie["runtime_min"],
                genres=tuple(raw_movie["genres"]),
                production_companies=tuple(raw_movie["production_companies"]),
                production_countries=tuple(raw_movie["production_countries"]),
                overview=str(raw_movie["overview"]),
                tagline=str(raw_movie["tagline"]),
                director=str(raw_movie["director"]),
                cast=tuple(raw_movie["cast"]),
                keywords=tuple(raw_movie["keywords"]),
                popularity=popularity,
                vote_average=vote_average,
                vote_count=vote_count,
                us_rating=str(raw_movie["us_rating"]),
                normalized_title=normalize_text(str(raw_movie["title"])),
                title_key=title_key(str(raw_movie["title"])),
                token_set=tokens,
                production_company_tokens=production_company_tokens,
                production_country_tokens=production_country_tokens,
                searchable_blob=normalize_text(joined),
                quality_score=quality_score,
            )
        )

    return tuple(movies)


@lru_cache(maxsize=1)
def movie_lookup() -> dict[int, Movie]:
    return {movie.tmdb_id: movie for movie in load_movies()}


@lru_cache(maxsize=1)
def title_lookup() -> dict[str, list[Movie]]:
    mapping: dict[str, list[Movie]] = {}
    for movie in load_movies():
        mapping.setdefault(movie.title_key, []).append(movie)
    return mapping


@lru_cache(maxsize=1)
def token_idf() -> dict[str, float]:
    movies = load_movies()
    doc_count = len(movies)
    counts: Counter[str] = Counter()
    for movie in movies:
        counts.update(movie.token_set)
    return {token: math.log((doc_count + 1) / (count + 1)) + 1.0 for token, count in counts.items()}


def build_semantic_text(movie: Movie) -> str:
    parts = [
        movie.title,
        " ".join(movie.genres),
        " ".join(movie.production_companies),
        " ".join(movie.production_countries),
        movie.overview,
        movie.tagline,
        movie.director,
        " ".join(movie.cast),
        " ".join(movie.keywords),
    ]
    return " | ".join(part for part in parts if part)


def _cosine_dense(left: list[float], right: list[float]) -> float:
    left_norm = math.sqrt(sum(value * value for value in left))
    right_norm = math.sqrt(sum(value * value for value in right))
    if left_norm == 0.0 or right_norm == 0.0:
        return 0.0
    dot = sum(a * b for a, b in zip(left, right))
    return dot / (left_norm * right_norm)


def _cosine_sparse(left: dict[str, float], right: dict[str, float]) -> float:
    if not left or not right:
        return 0.0
    left_norm = math.sqrt(sum(value * value for value in left.values()))
    right_norm = math.sqrt(sum(value * value for value in right.values()))
    if left_norm == 0.0 or right_norm == 0.0:
        return 0.0
    if len(left) > len(right):
        left, right = right, left
    dot = sum(value * right.get(token, 0.0) for token, value in left.items())
    return dot / (left_norm * right_norm)


def _try_sentence_transformer_backend(movie_texts: list[str]) -> dict[str, object] | None:
    try:
        sentence_transformers = importlib.import_module("sentence_transformers")
        sentence_transformer = getattr(sentence_transformers, "SentenceTransformer")
        model = sentence_transformer("all-MiniLM-L6-v2")
        embeddings = model.encode(
            movie_texts,
            normalize_embeddings=True,
            show_progress_bar=False,
        )
        return {
            "name": "sentence-transformers",
            "kind": "dense",
            "model": model,
            "movie_embeddings": [list(vector) for vector in embeddings],
        }
    except Exception:
        return None


def _build_tfidf_backend(movie_texts: list[str]) -> dict[str, object]:
    idf = token_idf()
    movie_embeddings = []
    for text in movie_texts:
        counts = Counter(tokenize(text))
        total_terms = sum(counts.values()) or 1
        movie_embeddings.append(
            {
                token: (count / total_terms) * idf.get(token, 0.0)
                for token, count in counts.items()
                if token in idf
            }
        )
    return {
        "name": "tfidf-fallback",
        "kind": "sparse",
        "idf": idf,
        "movie_embeddings": movie_embeddings,
    }


@lru_cache(maxsize=1)
def semantic_backend() -> dict[str, object]:
    movies = load_movies()
    movie_texts = [build_semantic_text(movie) for movie in movies]
    backend = _try_sentence_transformer_backend(movie_texts)
    if backend is None:
        backend = _build_tfidf_backend(movie_texts)
    backend["movie_embeddings_by_id"] = {
        movie.tmdb_id: backend["movie_embeddings"][index]
        for index, movie in enumerate(movies)
    }
    return backend


def build_semantic_query_text(
    preferences: str,
    query_weights: Counter[str],
    signals: dict[str, object],
) -> str:
    blocked_tokens = set(signals.get("excluded_tokens", set()))
    for genre in signals.get("avoided_genres", set()):
        blocked_tokens.update(tokenize(genre))

    positive_tokens = [
        token
        for token, weight in query_weights.most_common()
        if weight > 0 and token not in blocked_tokens
    ]

    parts = [preferences]
    preferred_genres = sorted(signals.get("preferred_genres", set()))
    if preferred_genres:
        parts.append("preferred genres " + " ".join(preferred_genres))

    agent_intent = signals.get("agent_intent", {})
    if isinstance(agent_intent, dict):
        vibe_keywords = [
            token
            for token in tokenize(" ".join(agent_intent.get("vibe_keywords", [])))
            if token not in blocked_tokens
        ]
        if vibe_keywords:
            parts.append("vibe " + " ".join(vibe_keywords[:10]))

    if positive_tokens:
        parts.append("focus " + " ".join(positive_tokens[:24]))
    return " ".join(part for part in parts if part)


def embed_semantic_query(
    preferences: str,
    query_weights: Counter[str],
    signals: dict[str, object],
) -> list[float] | dict[str, float]:
    backend = semantic_backend()
    query_text = build_semantic_query_text(preferences, query_weights, signals)
    if backend["kind"] == "dense":
        vector = backend["model"].encode(
            [query_text],
            normalize_embeddings=True,
            show_progress_bar=False,
        )[0]
        return list(vector)

    counts = Counter(tokenize(query_text))
    total_terms = sum(counts.values()) or 1
    idf = backend["idf"]
    return {
        token: (count / total_terms) * idf.get(token, 0.0)
        for token, count in counts.items()
        if token in idf
    }


def semantic_similarity(movie: Movie, query_embedding: list[float] | dict[str, float]) -> float:
    backend = semantic_backend()
    movie_embedding = backend["movie_embeddings_by_id"][movie.tmdb_id]
    if backend["kind"] == "dense":
        cosine = _cosine_dense(query_embedding, movie_embedding)
        return max(0.0, min(1.0, (cosine + 1.0) / 2.0))

    cosine = _cosine_sparse(query_embedding, movie_embedding)
    return max(0.0, min(1.0, cosine))


def heuristic_extract_preferences(preferences: str) -> dict[str, object]:
    normalized = normalize_text(preferences)
    tokens = tokenize(preferences)
    token_weights: Counter[str] = Counter()
    preferred_genres: set[str] = set()
    avoided_genres: set[str] = set()
    explicit_exclusions: set[str] = set()
    excluded_title_phrases: set[str] = set()
    preferred_languages: set[str] = set()
    preferred_company_tokens: set[str] = set()
    preferred_country_tokens: set[str] = set()
    preferred_keyword_phrases: set[str] = set()
    excluded_company_tokens: set[str] = set()
    prefer_non_english = False
    prefer_low_rated = False
    prefer_happy_ending = False
    min_year: int | None = None
    max_year: int | None = None
    min_runtime: int | None = None
    max_runtime: int | None = None
    strict_min_year = False
    strict_max_year = False
    strict_max_runtime = False

    for token in tokens:
        token_weights[token] += 2.0

    for phrase, hints in PHRASE_HINTS.items():
        phrase_norm = normalize_text(phrase)
        if contains_normalized_phrase(normalized, phrase_norm):
            is_negated = any(contains_normalized_phrase(normalized, neg.strip() + " " + phrase_norm) for neg in NEGATION_PHRASES)
            boost = -3.0 if is_negated else 2.5
            for hint in hints:
                for token in tokenize(hint):
                    token_weights[token] += boost

    for genre, aliases in GENRE_ALIASES.items():
        for alias in aliases | {genre}:
            alias_norm = normalize_text(alias)
            if alias_norm and contains_normalized_phrase(normalized, alias_norm):
                is_negated = False
                for neg in NEGATION_PHRASES:
                    idx = normalized.find(neg)
                    while idx >= 0:
                        window = normalized[idx: idx + 80]
                        if alias_norm in window:
                            is_negated = True
                            break
                        idx = normalized.find(neg, idx + 1)
                    if is_negated:
                        break
                if is_negated:
                    avoided_genres.add(genre)
                    for token in tokenize(alias):
                        explicit_exclusions.add(token)
                else:
                    preferred_genres.add(genre)
                    for token in tokenize(alias):
                        token_weights[token] += 3.5

    if any(phrase in normalized for phrase in ["non english", "non-english", "foreign language", "foreign", "international", "subtitled", "subtitles"]):
        prefer_non_english = True

    if "pixar" in normalized:
        preferred_genres.update({"animation", "family"})
        preferred_languages.add("en")
        for token in PIXAR_HINT_TOKENS:
            token_weights[token] += 3.5
    elif "disney" in normalized:
        preferred_genres.update({"animation", "family"})
        for token in PIXAR_HINT_TOKENS:
            token_weights[token] += 2.5

    if any(phrase in normalized for phrase in HAPPY_ENDING_AVOIDANCE_PHRASES):
        prefer_happy_ending = True
        for token in ["heartwarming", "hopeful", "uplifting", "friendship", "family", "fun"]:
            token_weights[token] += 2.5
        preferred_genres.update({"family", "comedy", "adventure"})

    if any(phrase in normalized for phrase in LOW_RATED_PHRASES):
        prefer_low_rated = True
    elif ("bad" in tokens or "terrible" in tokens or "awful" in tokens or "worst" in tokens) and ("movie" in normalized or "film" in normalized):
        prefer_low_rated = True
    elif "low" in tokens and ("rated" in tokens or "rating" in tokens):
        prefer_low_rated = True

    for language, codes in LANGUAGE_ALIASES.items():
        if contains_normalized_phrase(normalized, language):
            if language == "english" and prefer_non_english:
                continue
            preferred_languages.update(codes)

    for alias, hint in COUNTRY_ALIASES.items():
        if contains_normalized_phrase(normalized, alias):
            preferred_country_tokens.update(hint["country_tokens"])
            preferred_languages.update(hint["language_codes"])
            for token in hint["country_tokens"]:
                token_weights[token] += 3.0

    for alias, hint in STUDIO_ALIASES.items():
        if contains_normalized_phrase(normalized, alias):
            preferred_company_tokens.update(hint["company_tokens"])
            preferred_languages.update(hint["language_codes"])
            preferred_genres.update(hint["preferred_genres"])
            for token in hint["company_tokens"]:
                token_weights[token] += 4.0

    if any(contains_normalized_phrase(normalized, phrase) for phrase in TV_SERIES_REQUEST_PHRASES):
        preferred_keyword_phrases.update(TV_TIE_IN_KEYWORD_PHRASES)
        for token in ["series", "television", "adaptation"]:
            token_weights[token] += 3.0

    if any(word in normalized for word in ["older", "classic", "old school", "old-school", "vintage"]):
        max_year = 2005
    if any(word in normalized for word in ["recent", "new", "newer", "latest", "modern"]):
        min_year = 2015

    before_match = re.search(r"before\s+(19\d{2}|20\d{2})", normalized)
    if before_match:
        max_year = int(before_match.group(1))
        strict_max_year = True

    after_match = (
        re.search(r"after\s+(19\d{2}|20\d{2})", normalized)
        or re.search(r"since\s+(19\d{2}|20\d{2})", normalized)
        or re.search(r"post\s+(19\d{2}|20\d{2})", normalized)
        or re.search(r"\b(19\d{2}|20\d{2})\s+(?:or|and)\s+(?:later|newer|after)\b", normalized)
        or re.search(r"\b(19\d{2}|20\d{2})\s+onward(?:s)?\b", normalized)
    )
    if after_match:
        min_year = int(after_match.group(1))
        strict_min_year = True

    explicit_years = [int(match) for match in re.findall(r"\b(19\d{2}|20\d{2})\b", normalized)]
    if explicit_years and before_match is None and after_match is None:
        if "before" in normalized or "older" in normalized or "classic" in normalized:
            max_year = min(explicit_years)
        elif any(word in normalized for word in ["after", "recent", "later", "newer", "onward", "onwards", "since", "post"]):
            min_year = max(explicit_years)

    runtime_limit = extract_runtime_limit_minutes(preferences, normalized)
    if runtime_limit is not None:
        max_runtime = runtime_limit if max_runtime is None else min(max_runtime, runtime_limit)
        strict_max_runtime = True

    if "under two hours" in normalized or "under 2 hours" in normalized or "under 120" in normalized:
        max_runtime = 120
        strict_max_runtime = True
    if any(word in normalized for word in ["short", "quick", "tight"]):
        max_runtime = 115 if max_runtime is None else min(max_runtime, 115)
    if "epic" in normalized or any(word in normalized for word in ["sweeping", "long"]):
        min_runtime = 130

    for negation in NEGATION_PHRASES:
        start = 0
        while True:
            idx = normalized.find(negation, start)
            if idx < 0:
                break
            fragment = normalized[idx + len(negation) : idx + len(negation) + 60]
            cleaned_fragment = fragment.strip()
            if cleaned_fragment:
                title_phrase = title_key(cleaned_fragment)
                if len(title_phrase) >= 3:
                    excluded_title_phrases.add(title_phrase)
            fragment_norm = normalize_text(fragment)
            for alias, hint in EXCLUSION_HINTS.items():
                if contains_normalized_phrase(fragment_norm, alias):
                    excluded_company_tokens.update(hint["company_tokens"])
                    avoided_genres.update(hint["genre_tokens"])
                    explicit_exclusions.update(hint["company_tokens"])
                    explicit_exclusions.update(hint["genre_tokens"])
            for token in fragment.split()[:5]:
                if len(token) > 2 and token not in EXCLUSION_STOPWORDS:
                    explicit_exclusions.add(token)
            start = idx + len(negation)

    signals = {
        "normalized": normalized,
        "token_weights": token_weights,
        "preferred_genres": preferred_genres,
        "avoided_genres": avoided_genres,
        "excluded_tokens": explicit_exclusions,
        "excluded_title_phrases": excluded_title_phrases,
        "heuristic_eras": set(),
        "preferred_languages": preferred_languages,
        "preferred_company_tokens": preferred_company_tokens,
        "preferred_country_tokens": preferred_country_tokens,
        "preferred_keyword_phrases": preferred_keyword_phrases,
        "excluded_company_tokens": excluded_company_tokens,
        "prefer_non_english": prefer_non_english,
        "prefer_low_rated": prefer_low_rated,
        "prefer_happy_ending": prefer_happy_ending,
        "min_year": min_year,
        "max_year": max_year,
        "min_runtime": min_runtime,
        "max_runtime": max_runtime,
        "strict_min_year": strict_min_year,
        "strict_max_year": strict_max_year,
        "strict_max_runtime": strict_max_runtime,
    }

    # Heuristic Era Detection (multi-era support)
    if any(word in normalized for word in ["90s", "1990s", "nineties"]):
        signals["heuristic_eras"].add("90s")
    if any(word in normalized for word in ["2000s", "2000", "naughties"]):
        signals["heuristic_eras"].add("2000s")
    if any(word in normalized for word in ["classic", "old school", "golden age"]):
        signals["heuristic_eras"].add("classic")
    if any(word in normalized for word in ["recent", "newest", "last few years", "latest"]):
        signals["heuristic_eras"].add("recent")

    if not signals["heuristic_eras"]:
        signals["heuristic_eras"].add("any")

    return signals


def agentic_extract_intent(preferences: str) -> dict | None:
    api_key = os.getenv("OLLAMA_API_KEY")
    if not api_key:
        return None

    prompt = f"""
Analyze this movie request: "{preferences}"

Extract the user's intent into exactly this JSON format. No markdown or text.
{{
  "must_have_genres": [],
  "must_not_have_genres": [],
  "target_eras": ["any"], 
  "discovery_mode": "neutral", 
  "vibe_keywords": []
}}

Rules for values:
target_eras choices: list containing any of ["classic", "90s", "2000s", "recent", "any"]
discovery_mode choices: "hidden_gem" (unknown/underrated), "blockbuster" (popular/famous), "neutral"
vibe_keywords: max 5 english words describing pacing or vibe (e.g. short, fast, creepy)
"""

    try:
        response = requests.post(
            f"{OLLAMA_HOST}/api/chat",
            headers={"Authorization": f"Bearer {api_key}"},
            json={
                "model": MODEL,
                "messages": [
                    {
                        "role": "system",
                        "content": "You are an intent extractor. Return ONLY valid JSON."
                    },
                    {"role": "user", "content": prompt},
                ],
                "stream": False,
            },
            timeout=3,
        )
        response.raise_for_status()
        content = response.json().get("message", {}).get("content", "").strip()
        match = re.search(r"\{.*\}", content, re.DOTALL)
        if match:
            parsed = json.loads(match.group(0))
            return parsed
    except Exception:
        pass
    return None


def extract_preferences(preferences: str) -> dict[str, object]:
    signals = heuristic_extract_preferences(preferences)
    agent_signals = agentic_extract_intent(preferences)
    if agent_signals:
        signals["agent_intent"] = agent_signals
    return signals


def watched_movie_ids(history: list[str], history_ids: list[int]) -> set[int]:
    watched_ids = {int(movie_id) for movie_id in history_ids if isinstance(movie_id, int) or str(movie_id).isdigit()}
    for title in history:
        key = title_key(title)
        if key and key in title_lookup():
            watched_ids.update(movie.tmdb_id for movie in title_lookup()[key])
    return watched_ids


def build_history_profile(history: list[str], history_ids: list[int]) -> dict[str, Counter[str]]:
    lookup = movie_lookup()
    profile = {
        "genres": Counter(),
        "keywords": Counter(),
        "cast": Counter(),
        "director": Counter(),
        "tokens": Counter(),
    }

    for movie_id in watched_movie_ids(history, history_ids):
        movie = lookup.get(movie_id)
        if not movie:
            continue
        profile["genres"].update(normalize_text(genre) for genre in movie.genres)
        profile["keywords"].update(tokenize(" ".join(movie.keywords)))
        profile["cast"].update(tokenize(" ".join(movie.cast[:3])))
        profile["director"].update(tokenize(movie.director))
        profile["tokens"].update(movie.token_set)

    return profile


def build_query_weights(
    preferences: str, history: list[str], history_ids: list[int]
) -> tuple[Counter[str], dict[str, object], dict[str, Counter[str]]]:
    signals = extract_preferences(preferences)
    history_profile = build_history_profile(history, history_ids)
    query_weights: Counter[str] = Counter(signals["token_weights"])

    if "agent_intent" in signals:
        vibe_keywords = signals["agent_intent"].get("vibe_keywords", [])
        for vibe in vibe_keywords:
            for token in tokenize(vibe):
                query_weights[token] += 3.0

    history_strength = 1.0 if len(query_weights) < 4 and not signals["preferred_genres"] else 0.35

    for genre, count in history_profile["genres"].most_common(4):
        for token in tokenize(genre):
            query_weights[token] += history_strength * min(count, 2)

    for token, count in history_profile["keywords"].most_common(10):
        query_weights[token] += 0.35 * min(count, 3)

    for token, count in history_profile["cast"].most_common(6):
        query_weights[token] += 0.25 * min(count, 2)

    for token, count in history_profile["director"].most_common(3):
        query_weights[token] += 0.6 * min(count, 2)

    return query_weights, signals, history_profile


def score_movie(
    movie: Movie,
    query_weights: Counter[str],
    signals: dict[str, object],
    history_profile: dict[str, Counter[str]],
    watched_ids: set[int],
) -> float:
    if movie.tmdb_id in watched_ids:
        return float("-inf")

    idf = token_idf()
    score = 0.0

    for token, weight in query_weights.items():
        if token in movie.token_set:
            score += weight * idf.get(token, 1.0)

    movie_genres = {normalize_text(genre) for genre in movie.genres}
    prefer_low_rated = bool(signals.get("prefer_low_rated"))
    prefer_happy_ending = bool(signals.get("prefer_happy_ending"))

    if signals["preferred_languages"]:
        if movie.original_language in signals["preferred_languages"]:
            score += 10.0
        else:
            score -= 6.0

    if signals.get("preferred_country_tokens"):
        if movie.production_country_tokens & set(signals["preferred_country_tokens"]):
            score += 8.0
        else:
            score -= 4.0

    if signals.get("preferred_company_tokens"):
        if movie.production_company_tokens & set(signals["preferred_company_tokens"]):
            score += 10.0
        else:
            score -= 4.0

    preferred_keyword_phrases = set(signals.get("preferred_keyword_phrases", set()))
    if preferred_keyword_phrases:
        matched_keyword_phrases = sum(1 for phrase in preferred_keyword_phrases if phrase in movie.searchable_blob)
        if matched_keyword_phrases:
            score += 18.0 + (4.0 * (matched_keyword_phrases - 1))
        else:
            score -= 8.0

    if signals.get("excluded_company_tokens"):
        if movie.production_company_tokens & set(signals["excluded_company_tokens"]):
            score -= 30.0

    if signals["prefer_non_english"]:
        if movie.original_language and movie.original_language != "en":
            score += 16.0
        else:
            score -= 20.0

    min_year = signals["min_year"]
    max_year = signals["max_year"]
    if movie.year is not None:
        if min_year is not None:
            score += 5.0 if movie.year >= min_year else -18.0
        if max_year is not None:
            score += 5.0 if movie.year <= max_year else -28.0

    min_runtime = signals["min_runtime"]
    max_runtime = signals["max_runtime"]
    if movie.runtime_min is not None:
        if min_runtime is not None:
            score += 2.5 if movie.runtime_min >= min_runtime else -4.0
        if max_runtime is not None:
            score += 2.5 if movie.runtime_min <= max_runtime else -4.0

    if movie.year is not None and movie.year > 2025:
        score -= 8.0

    if prefer_happy_ending:
        for genre in ("family", "comedy", "adventure", "animation", "romance"):
            if genre in movie_genres:
                score += 2.5
        for genre in ("horror", "war", "history"):
            if genre in movie_genres:
                score -= 4.0

        happy_hits = sum(1 for token in HAPPY_THEME_TOKENS if token in movie.token_set)
        sad_hits = sum(1 for token in SAD_THEME_TOKENS if token in movie.token_set)
        score += min(happy_hits, 3) * 1.5
        score -= min(sad_hits, 3) * 2.5

    for genre in signals["preferred_genres"]:
        if genre in movie_genres:
            score += 8.0
        else:
            score -= 6.0

    for genre in signals["avoided_genres"]:
        if genre in movie_genres:
            score -= 25.0

    matched_preferred = sum(1 for genre in signals["preferred_genres"] if genre in movie_genres)
    preferred_count = len(signals["preferred_genres"])
    if preferred_count >= 2 and matched_preferred == preferred_count:
        score += 6.0
    elif preferred_count >= 2 and matched_preferred == 0:
        score -= 6.0

    for token in signals["excluded_tokens"]:
        if token in movie.token_set or token in movie.searchable_blob:
            score -= 18.0

    for title_phrase in signals.get("excluded_title_phrases", set()):
        if title_phrase and (title_phrase == movie.title_key or title_phrase in movie.title_key):
            score -= 1000.0

    shared_history_genres = sum(history_profile["genres"].get(genre, 0) for genre in movie_genres)
    score += min(shared_history_genres, 4) * 0.9

    keyword_overlap = sum(history_profile["keywords"].get(token, 0) for token in movie.token_set)
    score += min(keyword_overlap, 10) * 0.18

    if movie.director:
        director_overlap = sum(history_profile["director"].get(token, 0) for token in tokenize(movie.director))
        score += director_overlap * 0.8

    cast_overlap = sum(history_profile["cast"].get(token, 0) for token in tokenize(" ".join(movie.cast[:3])))
    score += min(cast_overlap, 3) * 0.5

    agent_intent = signals.get("agent_intent", {})
    if agent_intent:
        must_have = {normalize_text(g) for g in agent_intent.get("must_have_genres", [])}
        must_have.update(signals["preferred_genres"])
        
        must_not_have = {normalize_text(g) for g in agent_intent.get("must_not_have_genres", [])}
        must_not_have.update(signals["avoided_genres"])
        
        for g in must_have:
            if g and g not in movie_genres:
                score -= 100.0
        for g in must_not_have:
            if g and g in movie_genres:
                score -= 100.0

        target_eras = set(agent_intent.get("target_eras", []))
        if not target_eras or "any" in target_eras:
             # Fallback to heuristic era if agent failed
             h_eras = signals.get("heuristic_eras", {"any"})
             target_eras = h_eras
             
        if "any" not in target_eras and movie.year:
            match_any_era = False
            if "classic" in target_eras and movie.year < 1990: match_any_era = True
            if "90s" in target_eras and 1990 <= movie.year <= 1999: match_any_era = True
            if "2000s" in target_eras and 2000 <= movie.year <= 2014: match_any_era = True
            if "recent" in target_eras and movie.year >= 2015: match_any_era = True
            
            if match_any_era:
                score += 15.0
            else:
                score -= 100.0  # CRUSHING PENALTY

        discovery_mode = agent_intent.get("discovery_mode", "neutral")
        if prefer_low_rated:
            score -= movie.quality_score * 4.0
            score += (6.6 - movie.vote_average) * 10.0
        elif discovery_mode == "hidden_gem":
            score += (movie.vote_average - 7.0) * 4.0
            score -= min(movie.popularity / 50.0, 8.0) 
            score += movie.quality_score * 0.5 
        elif discovery_mode == "blockbuster":
            score += math.log1p(movie.popularity) * 2.0
            score += movie.quality_score * 3.0
        else:
            score += movie.quality_score * 3.0
    else:
        if prefer_low_rated:
            score -= movie.quality_score * 4.0
            score += (6.6 - movie.vote_average) * 10.0
        else:
            score += movie.quality_score * 3.0

    if prefer_low_rated:
        if movie.vote_average <= 5.5:
            score += 8.0
        elif movie.vote_average <= 6.2:
            score += 4.0
        elif movie.vote_average >= 8.0:
            score -= 12.0
        elif movie.vote_average >= 7.5:
            score -= 8.0
    else:
        if movie.vote_average < 6.0:
            score -= 6.0
        elif movie.vote_average < 6.5:
            score -= 2.5

        if movie.vote_average >= 7.5:
            score += 1.5

    if movie.vote_count >= 5000:
        score += 1.2

    if movie.us_rating == "R" and "family" in signals["normalized"]:
        score -= 8.0

    if "epic" in signals["normalized"] and movie.runtime_min is not None:
        if movie.runtime_min >= 130:
            score += 3.0
        elif movie.runtime_min <= 105:
            score -= 3.0

    return score


def explain_match(
    movie: Movie, preferences: str, signals: dict[str, object], history_profile: dict[str, Counter[str]]
) -> list[str]:
    reasons: list[str] = []
    movie_genres = {normalize_text(value) for value in movie.genres}

    if signals.get("prefer_low_rated"):
        if movie.vote_average <= 5.5:
            reasons.append("leans into the messy low-rated energy you're asking for")
        elif movie.vote_average <= 6.2:
            reasons.append("lands in the rough-around-the-edges zone you're asking for")

    if signals.get("prefer_happy_ending") and any(genre in movie_genres for genre in {"family", "comedy", "adventure", "animation"}):
        reasons.append("feels closer to the upbeat landing you're asking for")

    preferred_genres = [genre for genre in signals["preferred_genres"] if genre in movie_genres]
    if preferred_genres:
        reasons.append(f"leans into your taste for {', '.join(preferred_genres[:2])}")

    matched_keywords = [token for token, _ in signals["token_weights"].most_common(12) if token in movie.token_set]
    if matched_keywords:
        reasons.append(f"matches the {', '.join(matched_keywords[:3])} vibe you're asking for")

    history_genre_matches = [
        genre for genre, count in history_profile["genres"].most_common() if count and genre in movie_genres
    ]
    if history_genre_matches:
        reasons.append("fits the patterns in your watch history")

    if not signals.get("prefer_low_rated") and movie.vote_average >= 7.8:
        reasons.append("has the kind of quality that usually makes a pick feel worth it")

    return reasons[:3]


def enforce_description_limit(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    if len(cleaned) <= DESCRIPTION_LIMIT:
        return cleaned
    truncated = cleaned[: DESCRIPTION_LIMIT - 1].rstrip()
    if ". " in truncated:
        truncated = truncated.rsplit(". ", 1)[0].rstrip(". ")
    return truncated[: DESCRIPTION_LIMIT - 1].rstrip() + "..."


def deterministic_description(movie: Movie, preferences: str, history: list[str], reasons: list[str], start_time: float) -> str:
    movie_genres = {normalize_text(g) for g in movie.genres}
    pref_text = normalize_text(preferences)
    keyword_text = " ".join(movie.keywords).lower()
    low_rated_request = any(phrase in pref_text for phrase in LOW_RATED_PHRASES) or (
        ("bad" in pref_text or "terrible" in pref_text or "awful" in pref_text or "worst" in pref_text)
        and ("movie" in pref_text or "film" in pref_text)
    )
    style = choose_style_profile(low_rated_request)

    if low_rated_request:
        hook = f"{style['hook_prefix']} {movie.title} is a very watchable {movie.runtime_min}-minute pick."
    elif "animation" in movie_genres and any(word in pref_text for word in ["emotional", "heart", "animated", "warm"]):
        hook = f"{style['hook_prefix']} {movie.title} is a great {movie.runtime_min}-minute pick."
    elif "romance" in movie_genres and "comedy" in movie_genres:
        hook = f"{style['hook_prefix']} {movie.title} is a great {movie.runtime_min}-minute call."
    elif "horror" in movie_genres:
        hook = f"{style['hook_prefix']} {movie.title} is a smart and creepy {movie.runtime_min}-minute pick."
    elif "thriller" in movie_genres or "mystery" in movie_genres:
        hook = f"{style['hook_prefix']} {movie.title} is a sharp {movie.runtime_min}-minute choice."
    elif "science fiction" in movie_genres:
        hook = f"{style['hook_prefix']} {movie.title} is an imaginative {movie.runtime_min}-minute choice."
    else:
        hook = f"{style['hook_prefix']} {movie.title} is a strong {movie.runtime_min}-minute choice."

    fit_reasons = []

    if "animation" in movie_genres:
        fit_reasons.append("warmth")
    if "drama" in movie_genres:
        fit_reasons.append("emotional insight")
    if "comedy" in movie_genres:
        fit_reasons.append("personality")
    if "romance" in movie_genres:
        fit_reasons.append("chemistry")
    if "thriller" in movie_genres:
        fit_reasons.append("tension")
    if "mystery" in movie_genres:
        fit_reasons.append("intrigue")
    if "science fiction" in movie_genres:
        fit_reasons.append("an imaginative style")
    if "adventure" in movie_genres:
        fit_reasons.append("forward momentum")
    if "family" in movie_genres:
        fit_reasons.append("heart")

    if "music" in keyword_text or "jazz" in keyword_text:
        fit_reasons.append("a fresh creative energy")
    elif "friendship" in keyword_text or "family" in keyword_text:
        fit_reasons.append("a touching emotional core")
    elif "dream" in keyword_text or "reality" in keyword_text:
        fit_reasons.append("a memorable imaginative hook")
    elif movie.tagline:
        fit_reasons.append("a memorable hook")

    for reason in reasons:
        short_reason = reason.replace("leans into your taste for ", "").replace("matches the ", "")
        if short_reason and len(fit_reasons) < 4:
            fit_reasons.append(short_reason)

    deduped = []
    seen = set()
    for reason in fit_reasons:
        if reason not in seen:
            deduped.append(reason)
            seen.add(reason)

    fit_reasons = deduped[:3]

    if len(fit_reasons) >= 3:
        fit_sentence = (
            f"It {style['fit_verb']} {fit_reasons[0]}, {fit_reasons[1]}, and {fit_reasons[2]} "
            "in a way that feels vivid rather than generic."
        )
    elif len(fit_reasons) == 2:
        fit_sentence = f"It {style['fit_verb']} {fit_reasons[0]} and {fit_reasons[1]} in a way that feels vivid rather than generic."
    elif len(fit_reasons) == 1:
        fit_sentence = f"It leans into {fit_reasons[0]} in a way that feels specific and memorable."
    else:
        fit_sentence = "It has the kind of tone and payoff that makes it feel like a genuinely good pick."

    if "animation" in movie_genres and "drama" in movie_genres:
        payoff = f"{style['payoff_prefix']} it's the kind of movie that leaves you moved, lighter, and glad you picked it."
    elif "romance" in movie_genres and "comedy" in movie_genres:
        payoff = f"{style['payoff_prefix']} it feels breezy, appealing, and fun to spend time with from the start."
    elif "thriller" in movie_genres or "mystery" in movie_genres:
        payoff = f"{style['payoff_prefix']} it has the kind of payoff that makes the choice feel satisfying instead of obvious."
    elif "horror" in movie_genres:
        payoff = f"{style['payoff_prefix']} it feels intense in the right way and sticks with you after it ends."
    elif low_rated_request:
        payoff = f"{style['payoff_prefix']} it has the kind of flawed, chaotic energy that makes a bad-movie pick feel intentional rather than accidental."
    elif "science fiction" in movie_genres:
        payoff = f"{style['payoff_prefix']} it feels expansive and entertaining without losing the human side of the story."
    else:
        payoff = f"{style['payoff_prefix']} it feels like the kind of pick that lands well once you actually press play."

    description = f"{hook} {fit_sentence} {payoff}"
    return enforce_description_limit(description)


def remaining_budget_seconds(start_time: float) -> float:
    return REQUEST_TIMEOUT_SECONDS - (time.monotonic() - start_time)


def wants_low_rated_movie(preferences: str, signals: dict[str, object] | None = None) -> bool:
    normalized = normalize_text(preferences)
    if signals and signals.get("prefer_low_rated"):
        return True
    if any(phrase in normalized for phrase in LOW_RATED_PHRASES):
        return True
    tokens = set(tokenize(preferences))
    if {"bad", "terrible", "awful", "worst"} & tokens and ("movie" in normalized or "film" in normalized):
        return True
    if "low" in tokens and ("rated" in tokens or "rating" in tokens):
        return True
    return False


def wants_absolute_worst_movie(preferences: str) -> bool:
    normalized = normalize_text(preferences)
    if any(phrase in normalized for phrase in ABSOLUTE_WORST_PHRASES):
        return True
    if "worst" in normalized and "movie" in normalized and any(word in normalized for word in ["ever", "made", "all time", "all-time", "absolute", "single"]):
        return True
    if "lowest rated" in normalized or "lowest rating" in normalized:
        return True
    return False


def choose_style_profile(low_rated_mode: bool) -> dict[str, str]:
    profiles = LOW_RATED_STYLE_PROFILES if low_rated_mode else STYLE_PROFILES
    return random.choice(profiles)


def _post_json_chat(messages: list[dict[str, str]], timeout_seconds: float) -> dict | None:
    api_key = os.getenv("OLLAMA_API_KEY")
    if not api_key or timeout_seconds <= 0:
        return None

    try:
        response = requests.post(
            f"{OLLAMA_HOST}/api/chat",
            headers={"Authorization": f"Bearer {api_key}"},
            json={
                "model": MODEL,
                "messages": messages,
                "stream": False,
            },
            timeout=(2, max(1.0, timeout_seconds)),
        )
        response.raise_for_status()
        content = response.json().get("message", {}).get("content", "").strip()
        match = re.search(r"\{.*\}", content, re.DOTALL)
        if match:
            parsed = json.loads(match.group(0))
            if isinstance(parsed, dict):
                return parsed
    except Exception:
        return None
    return None


def critic_fix_instruction(
    movie: Movie,
    preferences: str,
    history: list[str],
    description: str,
    start_time: float,
) -> str | None:
    remaining_budget = remaining_budget_seconds(start_time)
    if remaining_budget < CRITIC_REVIEW_MIN_BUDGET_SECONDS:
        return None

    parsed = _post_json_chat(
        [
            {
                "role": "system",
                "content": (
                    "You are a skeptical classmate judging whether a movie recommendation blurb sounds specific, appealing, and tailored. "
                    "If the user explicitly wants a badly rated, messy, campy, or so-bad-it's-good movie, treat that as a valid target and judge the blurb on whether it embraces that flawed appeal rather than prestige or acclaim. "
                    "Return ONLY valid JSON with keys approve and fix_instruction. If the blurb is already strong, set approve to true and fix_instruction to an empty string. "
                    "If it is generic or weak, set approve to false and give exactly one concise sentence telling another agent how to improve it."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"User preferences: {preferences}\n"
                    f"Watch history: {', '.join(history[:8]) if history else 'None provided'}\n"
                    f"Chosen movie: {movie.title} ({movie.year or 'Unknown'})\n"
                    f"Runtime: {movie.runtime_min or 'Unknown'}\n"
                    f"Genres: {', '.join(movie.genres)}\n"
                    f"Overview: {movie.overview}\n"
                    f"Description to review: {description}"
                ),
            },
        ],
        min(CRITIC_REVIEW_TIMEOUT_SECONDS, remaining_budget),
    )
    if not parsed:
        return None
    if parsed.get("approve") is True:
        return None
    fix_instruction = str(parsed.get("fix_instruction", "")).strip()
    return fix_instruction or None


def rewrite_description_with_fix(
    movie: Movie,
    preferences: str,
    history: list[str],
    description: str,
    fix_instruction: str,
    start_time: float,
) -> str | None:
    remaining_budget = remaining_budget_seconds(start_time)
    if remaining_budget < CRITIC_REWRITE_MIN_BUDGET_SECONDS:
        return None

    parsed = _post_json_chat(
        [
            {
                "role": "system",
                "content": (
                    "You rewrite movie recommendation blurbs. Return ONLY valid JSON with a single key named description. "
                    "Keep it under 500 characters, mention the movie title in the first two sentences, avoid spoilers, and make it vivid rather than generic. "
                    "If the user wants a badly rated or so-bad-it's-good movie, lean into flawed, chaotic, campy, or messy appeal and do not apologize or pivot to acclaim."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Movie: {movie.title} ({movie.year or 'Unknown'})\n"
                    f"Runtime: {movie.runtime_min or 'Unknown'} minutes\n"
                    f"Genres: {', '.join(movie.genres)}\n"
                    f"Overview: {movie.overview}\n"
                    f"User preferences: {preferences}\n"
                    f"Watch history: {', '.join(history[:8]) if history else 'None provided'}\n"
                    f"Current description: {description}\n"
                    f"Fix instruction: {fix_instruction}\n"
                    f"Return JSON like {{\"description\": \"...\"}}"
                ),
            },
        ],
        min(CRITIC_REWRITE_TIMEOUT_SECONDS, remaining_budget),
    )
    if not parsed:
        return None
    rewritten = str(parsed.get("description", "")).strip()
    return enforce_description_limit(rewritten) if rewritten else None


def agentic_judge_and_describe(movies: list[Movie], preferences: str, history: list[str], signals: dict[str, object], history_profile: dict[str, Counter[str]], start_time: float) -> dict:
    fallback_movie = movies[0]
    fallback_reasons = explain_match(fallback_movie, preferences, signals, history_profile)
    fallback_desc = deterministic_description(fallback_movie, preferences, history, fallback_reasons, start_time)
    fallback_candidate = {"tmdb_id": fallback_movie.tmdb_id, "description": fallback_desc}

    if wants_absolute_worst_movie(preferences):
        return fallback_candidate

    api_key = os.getenv("OLLAMA_API_KEY")
    remaining_budget = remaining_budget_seconds(start_time)
    if not api_key or remaining_budget < 2.0:
        return fallback_candidate

    low_rated_mode = wants_low_rated_movie(preferences, signals)
    style = choose_style_profile(low_rated_mode)

    candidates_text = ""
    for i, m in enumerate(movies):
        candidates_text += f"[{i+1}] {m.title} ({m.year or 'Unknown'}) | Runtime: {m.runtime_min or 'Unknown'} min | tmdb_id: {m.tmdb_id}\nGenres: {', '.join(m.genres)}\nOverview: {m.overview}\n\n"

    request_mode_note = ""
    if low_rated_mode:
        request_mode_note = (
            "SPECIAL MODE - LOW-RATED REQUEST:\n"
            "The user explicitly wants a poorly rated, bad, messy, or so-bad-it's-good movie.\n"
            "Treat that as the goal, not as a mistake to correct.\n"
            "Do NOT apologize, do NOT claim the shortlist is critically acclaimed, and do NOT steer toward prestige or quality.\n"
            "Favor the lower-rated fitting option and describe its flawed, chaotic, campy, trashy, or hate-watch appeal in a positive, intentional way.\n\n"
        )

    style_note = (
        "WRITING STYLE DIRECTION:\n"
        f"{style['instruction']}\n"
        "This is a wording instruction only. Do not change the selected movie just to satisfy style.\n\n"
    )

    prompt = (
        "You are an expert, emotionally intelligent movie recommendation agent.\n"
        f"User preferences: {preferences}\n"
        f"Watch history: {', '.join(history[:8]) if history else 'None provided'}\n\n"
        f"{request_mode_note}"
        f"{style_note}"
        "Here are the top candidates that match their taste profile:\n"
        f"{candidates_text}"
        "Task:\n"
        "1. Act as a judge. Pick the single best candidate. Note: The list below is ALREADY RANKED by our mathematical engine. Respect the rank unless you find a reason it fails the user's specific tone/vibe.\n"
        f"2. Write a persuasive, emotionally resonant recommendation blurb (max {LLM_CHAR_BUDGET} chars, no spoilers, no bullet points). "
        "Follow these 6 rules for the blurb:\n"
        "   RULE 1 - OPEN WITH HISTORY: Start by referencing a specific movie from their Watch history that shares DNA with your pick (e.g. 'Since you loved [watched movie]...'). Skip if no history.\n"
        "   RULE 2 - WEAVE IN RUNTIME: In the first or second sentence, naturally embed the runtime (e.g. 'In this gripping [X]-minute thriller...'). Do NOT skip this.\n"
        "   RULE 3 - NAME THE MOVIE: You MUST explicitly use the title of the recommended movie in the first two sentences so the user knows exactly what you picked.\n"
        "   RULE 4 - VIVID DESCRIPTION: Use vivid, specific, emotionally charged language about the movie's vibe. Avoid generic words like 'great' or 'amazing'.\n"
        "   RULE 5 - DEFEND NEGATIVES: If the user expressed hates or avoidances, actively rebut them (e.g. 'This isn't a loud action flick, but a...'). This builds trust.\n"
        "   RULE 6 - COMPELLING CLOSE: End with one punchy sentence that makes them want to watch it right now.\n"
        "   RULE 7 - STRICT CONSTRAINTS: You MUST respect the Era (Year) and Avoided Genres. If the candidates list above has a movie that violates these, IGNORE IT. Do not recommend a 2017 movie if user asked for 90s/2000s.\n"
        "   RULE 8 - BAD-MOVIE REQUESTS: If the user asked for something low-rated or bad on purpose, do not frame the movie as acclaimed or prestigious. Sell the pick as intentionally flawed fun.\n"
        "3. Output ONLY a valid JSON object matching this exact shape:\n"
        '{"thought_process": "<why this movie perfectly matches in 15 words>", "tmdb_id": <selected tmdb_id integer>, "description": "<your blurb here>"}\n'
    )

    try:
        parsed = _post_json_chat(
            [
                {
                    "role": "system",
                    "content": "You are a movie recommendation AI. You MUST output ONLY valid JSON without markdown wrapping."
                },
                {"role": "user", "content": prompt},
            ],
            min(remaining_budget, REQUEST_TIMEOUT_SECONDS),
        )
        if parsed and "tmdb_id" in parsed and "description" in parsed:
            candidate = {
                "tmdb_id": int(parsed["tmdb_id"]),
                "description": enforce_description_limit(str(parsed["description"]))
            }
            chosen_movie = next((movie for movie in movies if movie.tmdb_id == candidate["tmdb_id"]), None)
            if chosen_movie is None:
                return candidate

            fix_instruction = critic_fix_instruction(
                chosen_movie,
                preferences,
                history,
                candidate["description"],
                start_time,
            )
            if not fix_instruction:
                return candidate

            rewritten = rewrite_description_with_fix(
                chosen_movie,
                preferences,
                history,
                candidate["description"],
                fix_instruction,
                start_time,
            )
            if rewritten:
                candidate["description"] = rewritten
            return candidate
        raise ValueError("Invalid JSON schema returned by LLM.")
    except Exception:
        return fallback_candidate


def validate_output(candidate: dict[str, object], watched_ids: set[int]) -> dict[str, object]:
    movie_id = candidate.get("tmdb_id")
    description = str(candidate.get("description", "")).strip()

    if not isinstance(movie_id, int):
        raise ValueError("tmdb_id must be an int")
    if movie_id not in movie_lookup():
        raise ValueError("tmdb_id not in candidate list")
    if movie_id in watched_ids:
        raise ValueError("recommended movie is already watched")
    if not description:
        raise ValueError("description cannot be empty")

    return {
        "tmdb_id": movie_id,
        "description": enforce_description_limit(description),
    }


def choose_top_movies(
    preferences: str, history: list[str], history_ids: list[int], top_k: int = 5
) -> tuple[list[Movie], dict[str, object], dict[str, Counter[str]], set[int]]:
    watched_ids = watched_movie_ids(history, history_ids)
    query_weights, signals, history_profile = build_query_weights(preferences, history, history_ids)
    strict_min_year = bool(signals.get("strict_min_year"))
    strict_max_year = bool(signals.get("strict_max_year"))
    min_year = signals.get("min_year")
    max_year = signals.get("max_year")
    strict_max_runtime = bool(signals.get("strict_max_runtime"))
    max_runtime = signals.get("max_runtime")

    base_scored = []
    eligible_movies = []
    for movie in load_movies():
        if strict_min_year:
            if movie.year is None or min_year is None or movie.year < min_year:
                continue
        if strict_max_year:
            if movie.year is None or max_year is None or movie.year > max_year:
                continue
        if strict_max_runtime:
            if movie.runtime_min is None or max_runtime is None or movie.runtime_min > max_runtime:
                continue
        eligible_movies.append(movie)
        score = score_movie(movie, query_weights, signals, history_profile, watched_ids)
        if math.isfinite(score):
            base_scored.append((score, movie))

    base_scored.sort(key=lambda item: item[0], reverse=True)

    rerank_pool = base_scored[: max(top_k, SEMANTIC_RERANK_POOL_SIZE)]
    query_embedding = embed_semantic_query(preferences, query_weights, signals)
    reranked = []
    for base_score, movie in rerank_pool:
        semantic_score = semantic_similarity(movie, query_embedding)
        final_score = base_score + (SEMANTIC_SCORE_WEIGHT * semantic_score)
        reranked.append((final_score, base_score, movie))

    reranked.sort(key=lambda item: (item[0], item[1]), reverse=True)
    top_movies = [movie for final_score, base_score, movie in reranked[:top_k]]

    if not top_movies:
        fallback_pool = [movie for movie in eligible_movies if movie.tmdb_id not in watched_ids]
        if not fallback_pool:
            fallback_pool = [movie for movie in load_movies() if movie.tmdb_id not in watched_ids]
        fallback = max(fallback_pool, key=lambda item: item.quality_score)
        top_movies = [fallback]

    return top_movies, signals, history_profile, watched_ids


def get_recommendation(preferences: str, history: list[str], history_ids: list[int] = []) -> dict:
    start_time = time.monotonic()
    movies, signals, history_profile, watched_ids = choose_top_movies(preferences, history, history_ids, top_k=5)
    candidate = agentic_judge_and_describe(movies, preferences, history, signals, history_profile, start_time)
    return validate_output(candidate, watched_ids)


def _parse_history_arg(values: Iterable[str] | None) -> list[str]:
    if not values:
        return []
    items: list[str] = []
    for value in values:
        for part in str(value).split("|"):
            cleaned = part.strip()
            if cleaned:
                items.append(cleaned)
    return items


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a movie recommendation from the BAMS 521 candidate set.")
    parser.add_argument("--preferences", help="Free-text user preferences.")
    parser.add_argument("--history", nargs="*", help="Watched movie titles. Use repeated args or pipe-separated values.")
    parser.add_argument("--history-ids", nargs="*", type=int, help="TMDB IDs corresponding to the watch history.")
    args = parser.parse_args()

    preferences = args.preferences or input("Preferences: ").strip()
    history = (
        _parse_history_arg(args.history)
        if args.history is not None
        else _parse_history_arg([input("Watch history (optional): ").strip()])
    )
    history_ids = args.history_ids or []
    print(get_recommendation(preferences, history, history_ids))


if __name__ == "__main__":
    main()
